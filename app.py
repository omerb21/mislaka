import json
import os
import logging
from datetime import datetime
from io import BytesIO

import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, send_file, flash, session
from werkzeug.utils import secure_filename

from process_pensions import (
    PensionFileProcessor,
    SEVERANCE_COLUMN_TAGS,
    TAGMUL_PERIOD_COLUMNS,
)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Change this to a secure secret key
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['PROCESSED_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'processed')

# Ensure storage folders exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['PROCESSED_FOLDER'], exist_ok=True)

BASE_COLUMNS = [
    'מספר חשבון',
    'שם תכנית',
    'חברה מנהלת',
    'יתרה',
    'תאריך התחלה',
    'סוג מוצר',
]

SEVERANCE_COLUMNS = list(SEVERANCE_COLUMN_TAGS.keys())
TAGMUL_COLUMNS = list(TAGMUL_PERIOD_COLUMNS.values())

TAIL_COLUMNS = [
    'סך תגמולים',
    'סך פיצויים',
    'סך רכיבים',
    'פער יתרה מול רכיבים',
    'תאריך נכונות יתרה',
    'מעסיקים היסטוריים',
]

TABLE_COLUMNS = BASE_COLUMNS + SEVERANCE_COLUMNS + TAGMUL_COLUMNS + TAIL_COLUMNS

NUMERIC_COLUMNS = {
    'יתרה',
    'סך תגמולים',
    'סך פיצויים',
    'סך רכיבים',
    'פער יתרה מול רכיבים',
    *SEVERANCE_COLUMNS,
    *TAGMUL_COLUMNS,
}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xml'}


def process_pension_file(filepath):
    """Process a single pension file and return the structured result."""
    try:
        processor = PensionFileProcessor(filepath)
        return processor.process()
    except Exception as e:
        logging.error(f"Error processing {filepath}: {str(e)}")
        return None


def _as_float(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def flatten_accounts(result):
    flattened = []
    if not result:
        return flattened

    for account in result.get('accounts', []):
        row = {
            'מספר חשבון': account.get('מספר_חשבון', ''),
            'שם תכנית': account.get('שם_תכנית', ''),
            'חברה מנהלת': account.get('חברה_מנהלת', ''),
            'יתרה': account.get('יתרה', 0.0),
            'תאריך התחלה': account.get('תאריך_התחלה', ''),
            'סוג מוצר': account.get('סוג_מוצר', ''),
        }

        severance_components = account.get('רכיבי_פיצויים', {}) or {}
        for column_name in SEVERANCE_COLUMNS:
            row[column_name] = severance_components.get(column_name, 0.0)

        tagmul_periods = account.get('תגמולים_לפי_תקופה', {}) or {}
        for column_name in TAGMUL_COLUMNS:
            row[column_name] = tagmul_periods.get(column_name, 0.0)

        tail_values = {
            'סך תגמולים': account.get('סך_תגמולים', 0.0),
            'סך פיצויים': account.get('סך_פיצויים', 0.0),
            'סך רכיבים': account.get('סך_רכיבים', 0.0),
            'פער יתרה מול רכיבים': account.get('פער_יתרה_מול_רכיבים', 0.0),
            'תאריך נכונות יתרה': account.get('תאריך_נכונות_יתרה', ''),
            'מעסיקים היסטוריים': account.get('מעסיקים_היסטוריים', ''),
        }

        for column_name in TAIL_COLUMNS:
            value = tail_values.get(column_name, '')
            row[column_name] = value

        flattened.append(row)

    return flattened


@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('לא נבחר קובץ', 'error')
            return redirect(request.url)
        
        files = request.files.getlist('file')
        if not files or files[0].filename == '':
            flash('לא נבחר קובץ', 'error')
            return redirect(request.url)
        
        # Remove previous processed file reference if exists
        previous_file = session.pop('results_file', None)
        if previous_file:
            old_path = os.path.join(app.config['PROCESSED_FOLDER'], previous_file)
            if os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except OSError:
                    logging.warning(f"Unable to remove old processed file: {old_path}")

        # Process each file
        all_rows = []
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)

                # Process the pension file
                try:
                    result = process_pension_file(filepath)
                    if result:
                        all_rows.extend(flatten_accounts(result))
                except Exception as e:
                    flash(f'שגיאה בעיבוד הקובץ {filename}: {str(e)}', 'error')
                    continue
            else:
                flash(f'סוג קובץ לא חוקי: {file.filename}', 'error')

        if all_rows:
            # Persist processed data to disk and keep reference in session
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            processed_filename = f'processed_{timestamp}.json'
            processed_path = os.path.join(app.config['PROCESSED_FOLDER'], processed_filename)

            with open(processed_path, 'w', encoding='utf-8') as processed_file:
                json.dump(all_rows, processed_file, ensure_ascii=False)

            session['results_file'] = processed_filename

            # Prepare DataFrame for display
            df = pd.DataFrame(all_rows, columns=TABLE_COLUMNS)
            numeric_cols = [col for col in df.columns if col in NUMERIC_COLUMNS]
            df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors='coerce')

            display_df = df.copy()
            for col in numeric_cols:
                display_df[col] = display_df[col].apply(
                    lambda x: f'{x:,.2f}' if pd.notnull(x) else ''
                )

            df_columns = display_df.columns.tolist()
            df_values = display_df.fillna('').values.tolist()
            numeric_column_indexes = [
                idx for idx, col in enumerate(df_columns) if col in numeric_cols
            ]

            totals_map = {col: df[col].sum(min_count=1) for col in numeric_cols}
            totals_row = []
            for idx, col in enumerate(df_columns):
                if idx == 0:
                    totals_row.append('סה"כ')
                elif col in totals_map and pd.notnull(totals_map[col]):
                    totals_row.append(f'{totals_map[col]:,.2f}')
                elif col in totals_map:
                    totals_row.append('0.00')
                else:
                    totals_row.append('')

            return render_template(
                'results.html',
                df_columns=df_columns,
                df_values=df_values,
                totals_row=totals_row,
                record_count=len(df),
                timestamp=datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                numeric_columns=numeric_cols,
                numeric_column_indexes=numeric_column_indexes,
                wide_layout=True,
            )
        else:
            flash('לא בוצע עיבוד של קבצים', 'error')
            return redirect(request.url)

    return render_template('upload.html')


@app.route('/export')
def export():
    processed_filename = session.get('results_file')
    if not processed_filename:
        flash('אין נתונים לייצוא', 'error')
        return redirect(url_for('upload_file'))

    processed_path = os.path.join(app.config['PROCESSED_FOLDER'], processed_filename)
    if not os.path.exists(processed_path):
        flash('קובץ העיבוד לא נמצא. אנא עבד מחדש את הקבצים.', 'error')
        session.pop('results_file', None)
        return redirect(url_for('upload_file'))

    try:
        with open(processed_path, 'r', encoding='utf-8') as processed_file:
            data = json.load(processed_file)

        df = pd.DataFrame(data, columns=TABLE_COLUMNS)
        numeric_cols = [col for col in df.columns if col in NUMERIC_COLUMNS]
        df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors='coerce')

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='נתוני פנסיה')
            workbook = writer.book
            worksheet = writer.sheets['נתוני פנסיה']

            header_format = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_format
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='right', vertical='center')

            number_format = '#,##0.00'
            for col_num, column_title in enumerate(df.columns, 1):
                column_letter = get_column_letter(col_num)
                max_length = len(str(column_title))
                for cell in worksheet[column_letter]:
                    if cell.row == 1:
                        continue
                    value = cell.value
                    if column_title in NUMERIC_COLUMNS and isinstance(value, (int, float)):
                        cell.number_format = number_format
                    if value is not None:
                        max_length = max(max_length, len(str(value)))
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

        output.seek(0)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'pension_results_{timestamp}.xlsx'
        )

    except Exception as e:
        logging.error(f"Export error: {str(e)}", exc_info=True)
        flash(f'שגיאה ביצוא הקובץ: {str(e)}', 'error')
        return redirect(url_for('upload_file'))

if __name__ == '__main__':
    app.run(debug=True)
